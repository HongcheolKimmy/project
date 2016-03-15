from openpyxl import load_workbook
from openpyxl import Workbook
import xlrd
from sets import Set
import sys

reload(sys)
sys.setdefaultencoding('utf-8')


class QCscanner:

    def __init__(self, filename):

        self.excelsheet = filename
        self.excelsheet = load_workbook(filename=self.excelsheet)

    def excel_gettabs(self, tabname, indicator = 0):
        encoded = self.excelsheet.get_sheet_names()
        tablist = []
        if tabname == "":
            for l in encoded:
                tablist.append(l)
        else:
            for i in encoded:
                if tabname in i:
                    tablist.append(i)
                else:
                    pass
        if indicator == 1:
            print tablist
        return tablist

    def excel_getdata(self, worksheet, text, indicator = 0):
        data = self.excelsheet.get_sheet_by_name(worksheet)
        cell = data.rows
        for i in cell:
            for l in i:
                datavalue = str(l.value)
                if text in datavalue:
                    num = str(l)
                    number = num.find(".") + 1
                    loc = str(l)[number:-1]
                    result = datavalue + "\t" + "@@" + loc
                    if indicator == 1:
                        print result
                    return result
                else:
                    continue

    def excel_getperiod(self,worksheet):
        periods = Set([])
        encoded = self.excelsheet.get_sheet_names()
        for i in encoded:
            if worksheet in i:
                sheet_name = i
            else:
                pass
        data = self.excelsheet.get_sheet_by_name(sheet_name)
        cell = data.range('A1:Z1')
        for i in cell:
            for l in i:
                if len(str(l.value)) >= 16:
                    periods.add(str(l.value)[:4])
                else:
                    pass
        for a in periods:
            print a,

    def excel_get_all_periods(self,worksheet,indicator = 0):
        encoded = self.excelsheet.get_sheet_names()
        for i in encoded:
            if worksheet in i:
                sheet_name = i
            else:
                pass
        data = self.excelsheet.get_sheet_by_name(sheet_name)
        cell = data.range('A1:Z1')
        result = []
        for i in cell:
            for l in i:
                if len(str(l.value)) >= 16:
                    num = str(l).find(".")
                    halfresult = l.value + "@@" + str(l)[num+1:num+2]
                    result.append(str(halfresult))
                else:
                    pass
        if indicator == 1:
            for a in result:
                print a
        return result

    def excel_getvalue(self, worksheet, field, period, indicator=0):
        valueloc = []
        yloc = self.excel_get_all_periods(worksheet)
        for i in yloc:
            if period in i:
                yloc_decodeloc = str(i).find("@@")
                ylock_decode = str(i)[int(yloc_decodeloc)+2:]
                valueloc.append(ylock_decode)
        xloc = self.excel_getdata(worksheet, field)
        xloc_decodeloc = str(xloc).find("@@")
        xloc_decode = str(xloc)[int(xloc_decodeloc)+3:]
        valueloc.append(xloc_decode)
        result = "".join(valueloc)
        valueloc_decode = result
        value = self.excelsheet.get_sheet_by_name(worksheet).cell(valueloc_decode)
        if indicator == 1:
            print value.value
        return value.value

    def QC_getdata(self, worksheet, text, indicator = 0):
        data = self.excelsheet.get_sheet_by_name(worksheet)
        cell = data.rows
        for i in cell:
            for l in i:
                datavalue = str(l.value)
                if text in datavalue:
                    num = str(l)
                    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                    number = num.find(".") + 1
                    loc = str(l)[number:-1]
                    newloc1 = alphabet.find(loc[:1]) + 2
                    newnewloc1 = str(alphabet[newloc1]) +loc[1:]
                    newloc2 = alphabet.find(loc[:1]) + 3
                    newnewloc2 = str(alphabet[newloc2]) + loc[1:]
                    result = datavalue + "\t" + "@@" + loc + " &&" + newnewloc1 + newnewloc2
                    if indicator == 1:
                        print result
                    return result
                else:
                    continue
    def QC_getvalue(self, worksheet, text, indicator = 0):
        data = self.excelsheet.get_sheet_by_name(worksheet)
        cell = data.rows
        result = []
        for i in cell:
            for l in i:
                datavalue = str(l.value)
                if text in datavalue:
                    num = str(l)
                    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
                    number = num.find(".") + 1
                    loc = str(l)[number:-1]
                    newloc1 = alphabet.find(loc[:1]) + 2
                    newnewloc1 = str(alphabet[newloc1]) +loc[1:]
                    newloc2 = alphabet.find(loc[:1]) + 3
                    newnewloc2 = str(alphabet[newloc2]) + loc[1:]
                    result.append(str(data.cell(newnewloc1).value))
                    result.append(str(data.cell(newnewloc2).value))
                    if indicator == 1:
                        print result
                    return result
                else:
                    continue




test = QCscanner('test111.xlsx')
test.excel_gettabs("",1)
# test.excel_getdata("FLEX - Income Statement-CORE", "IS064", 1)
# test.excel_getperiod("FLEX - Income Statement-CORE")
# test.excel_get_all_periods("FLEX - Income Statement-CORE", 1)
# test.excel_getvalue("FLEX - Income Statement-CORE", "IS097", "2013-A1-M(FP)-IA",1)
# test.excel_getvalue("Sheet 1", "A1041", "2013-A1-M(FP)-IA")ome Statement-ARD","S3862")
# test2 = QCscanner('test4.xlsx')
# test2.excel_gettabs("")
# test2.QC_getdata("IS", "IM177",1)
# test2.QC_getvalue("IS", "IM125",1)
