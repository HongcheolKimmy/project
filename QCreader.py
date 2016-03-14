from openpyxl import load_workbook
from sets import Set
import sys

reload(sys)
sys.setdefaultencoding('utf-8')


class QCscanner:

    def __init__(self, filename):
        self.excelsheet = filename
        self.excelsheet = load_workbook(filename=self.excelsheet)

    def excel_gettabs(self, tabname):
        encoded = self.excelsheet.get_sheet_names()
        if tabname == "":
            for l in encoded:
                print unicode(l)
        else:
            for i in encoded:
                if tabname in i:
                    print unicode(i)
                else:
                    pass

    def excel_getdata(self, worksheet, text):
        data = self.excelsheet.get_sheet_by_name(worksheet)
        cell = data.rows
        for i in cell:
            for l in i:
                datavalue = str(l.value)
                if text in datavalue:
                    num = str(l)
                    number = num.find(".") + 1
                    loc = str(l)[number:-1]
                    result = datavalue + "\t" + "@@" + loc
                    return result
                else:
                    pass

    def excel_getperiod(self,worksheet):
        periods = Set([])
        encoded = self.excelsheet.get_sheet_names()
        for i in encoded:
            if worksheet in i:
                sheet_name = i
            else:
                pass
        data = self.excelsheet.get_sheet_by_name(sheet_name)
        cell = data.iter_rows('A2:Z2')
        for i in cell:
            for l in i:
                if len(str(l.value)) >= 16:
                    periods.add(str(l.value)[:4])
                else:
                    pass
        print periods

    def excel_get_all_periods(self,worksheet):
        encoded = self.excelsheet.get_sheet_names()
        for i in encoded:
            if worksheet in i:
                sheet_name = i
            else:
                pass
        data = self.excelsheet.get_sheet_by_name(sheet_name)
        cell = data.iter_rows('A2:Z2')
        result = []
        for i in cell:
            for l in i:
                if len(str(l.value)) >= 16:
                    num = str(l).find(".")
                    halfresult = l.value + "@@" + str(l)[num+1:num+2]
                    result.append(str(halfresult))
                else:
                    pass
        return result

    def excel_getvalue(self, worksheet, field, period):
        valueloc = []
        yloc = self.excel_get_all_periods(worksheet)
        for i in yloc:
            if period in i:
                yloc_decodeloc = str(i).find("@@")
                ylock_decode = str(i)[int(yloc_decodeloc)+2:]
                valueloc.append(ylock_decode)
        xloc = self.excel_getdata(worksheet, field)
        xloc_decodeloc = str(xloc).find("@@")
        xloc_decode = str(xloc)[int(xloc_decodeloc)+3:]
        valueloc.append(xloc_decode)
        result = "".join(valueloc)
        valueloc_decode = result
        value = self.excelsheet.get_sheet_by_name(worksheet).cell(valueloc_decode)
        print value.value
        return value.value




test = QCscanner('test.xlsx')
test.excel_gettabs("Sheet")
test.excel_getdata("Sheet 1", "EPS")
test.excel_getperiod("Sheet 1")
test.excel_get_all_periods("Sheet 1")
test.excel_getvalue("Sheet 1", "A1011", "2013-A1-M(FP)-IA")
test.excel_getvalue("Sheet 1", "A1041", "2013-A1-M(FP)-IA")